#!/usr/bin/env python3
"""
nested_json_to_excel.py

Convert JSON nested hasil receipt_parser.py (list of customer/loan, tiap
customer punya field customer/loan di level luar + array "installments")
menjadi satu file .xlsx berisi 3 sheet:

  1. Flattened     - satu baris per installment, field customer/loan diulang
                      di tiap baris (pendekatan denormalized, paling gampang
                      dibuka & difilter langsung di Excel).
  2. Customers      - satu baris per loan_id, hanya field customer/loan.
  3. Installments   - satu baris per installment + kolom loan_id (foreign key)
                      untuk dihubungkan balik ke sheet Customers (mis. lewat
                      VLOOKUP/INDEX-MATCH atau Power Query di Excel).

Tidak ada kalkulasi/formula di sini (murni konversi struktur data), jadi tidak
perlu proses recalc.

Cara pakai:
    python nested_json_to_excel.py -i nested_output.json -o output.xlsx
"""

import argparse
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FONT = Font(name=FONT_NAME, bold=True)
BODY_FONT = Font(name=FONT_NAME)

# Banded rows: baris genap abu-abu muda, baris ganjil putih (no fill)
BAND_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Field customer/loan yang ada di level luar tiap objek (lihat receipt_parser.py)
CUSTOMER_LEVEL_FIELDS = [
    "loan_id", "customer_name", "email", "phone_number", "address", "city", "state",
    "postal_code", "loan_purpose", "loan_amount", "total_payable",
    "loan_duration_months", "interest_rate",
]

# Field per installment (di dalam array "installments")
INSTALLMENT_LEVEL_FIELDS = [
    "reference_number", "payment_date", "installment_number", "total_installments",
    "payment_method", "payment_amount", "total_paid_to_date", "remaining_balance",
    "current_repayment_status", "filename", "pageNumber",
]


def load_nested_json(path: str) -> list:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError("JSON input harus berupa list of customer/loan (hasil build_nested_records).")
    return data


def build_rows(customers: list):
    """Dari list nested, hasilkan 3 kumpulan baris: flattened, customers, installments."""
    flattened_rows = []
    customer_rows = []
    installment_rows = []

    for customer in customers:
        customer_row = {field: customer.get(field) for field in CUSTOMER_LEVEL_FIELDS}
        customer_rows.append(customer_row)

        installments = customer.get("installments", []) or [{}]  # tetap muncul 1 baris walau kosong
        for inst in installments:
            inst_row = {field: inst.get(field) for field in INSTALLMENT_LEVEL_FIELDS}
            # loan_id sebagai foreign key di sheet Installments
            inst_row_with_fk = {"loan_id": customer.get("loan_id"), **inst_row}
            installment_rows.append(inst_row_with_fk)

            # Flattened: gabungan customer fields + installment fields dalam satu baris
            flat_row = {**customer_row, **inst_row}
            flattened_rows.append(flat_row)

    return flattened_rows, customer_rows, installment_rows


def write_sheet(ws, rows: list, headers: list):
    """Tulis header (bold) lalu semua baris data ke worksheet, dan auto-size lebar kolom."""
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = HEADER_FONT

    for row in rows:
        ws.append([row.get(h) for h in headers])

    for row_offset, row in enumerate(ws.iter_rows(min_row=2)):
        # row_offset=0 -> baris data ke-1 (baris sheet ke-2) dst.
        is_banded_row = row_offset % 2 == 1
        for cell in row:
            cell.font = BODY_FONT
            if is_banded_row:
                cell.fill = BAND_FILL

    ws.freeze_panes = "A2"

    # Auto-size kolom berdasarkan panjang konten terpanjang (dibatasi biar tidak kelewat lebar)
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            val = row.get(header)
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)


def convert(input_path: str, output_path: str):
    customers = load_nested_json(input_path)
    flattened_rows, customer_rows, installment_rows = build_rows(customers)

    wb = Workbook()

    ws_flat = wb.active
    ws_flat.title = "Flattened"
    flat_headers = CUSTOMER_LEVEL_FIELDS + INSTALLMENT_LEVEL_FIELDS
    write_sheet(ws_flat, flattened_rows, flat_headers)

    ws_customers = wb.create_sheet("Customers")
    write_sheet(ws_customers, customer_rows, CUSTOMER_LEVEL_FIELDS)

    ws_installments = wb.create_sheet("Installments")
    installment_headers = ["loan_id"] + INSTALLMENT_LEVEL_FIELDS
    write_sheet(ws_installments, installment_rows, installment_headers)

    wb.save(output_path)
    return len(flattened_rows), len(customer_rows), len(installment_rows)


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convert JSON nested (hasil receipt_parser.py) menjadi Excel 3 sheet: "
                    "Flattened, Customers, Installments."
    )
    parser.add_argument("-i", "--input", required=True, dest="input_path",
                         help="Path file JSON nested input (hasil build_nested_records)")
    parser.add_argument("-o", "--output", required=True, dest="output_path",
                         help="Path file .xlsx output")
    return parser


def main():
    args = build_arg_parser().parse_args()
    try:
        n_flat, n_cust, n_inst = convert(args.input_path, args.output_path)
    except FileNotFoundError:
        print(f"Error: file input tidak ditemukan -> {args.input_path}", file=sys.stderr)
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"Error: file input bukan JSON valid -> {e}", file=sys.stderr)
        sys.exit(1)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"Selesai. {args.output_path} dibuat dengan:")
    print(f"  - Flattened   : {n_flat} baris")
    print(f"  - Customers   : {n_cust} baris")
    print(f"  - Installments: {n_inst} baris")


if __name__ == "__main__":
    main()